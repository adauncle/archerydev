# -*- coding: utf-8 -*-
"""把 v3 Excel 汇总统计/按风险等级/按负责人 公式范围从 33 扩到 100"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from openpyxl import load_workbook

path = r"G:\MiniMax工作空间\archery_dev\docs\reports\2026-08-06_功能开发计划_v3_172831.xlsx"
wb = load_workbook(path)

# 要改的 sheet 和公式范围: 33 → 100
SHEETS = ["汇总统计", "按风险等级", "按负责人"]
OLD_ROWS = "2:33"
NEW_ROWS = "2:100"
NEW_MAX = "100"

replacements = 0
for sn in SHEETS:
    ws = wb[sn]
    for row in ws.iter_rows():
        for c in row:
            if c.value and isinstance(c.value, str) and c.value.startswith("="):
                new_v = c.value
                import re
                # 跨 sheet 引用: 功能开发计划!列字母+数字:列字母+33 → ...:100
                new_v = re.sub(
                    r"(\b功能开发计划!)([A-Z]+)(\d+):([A-Z]+)33\b",
                    lambda m: f"{m.group(1)}{m.group(2)}{m.group(3)}:{m.group(4)}{NEW_MAX}",
                    new_v,
                )
                # 兜底: 本 sheet 内 A2:A33 → A2:A100
                new_v = re.sub(
                    r"\b([A-Z]+)(\d+):([A-Z]+)33\b",
                    lambda m: f"{m.group(1)}{m.group(2)}:{m.group(3)}{NEW_MAX}",
                    new_v,
                )
                if new_v != c.value:
                    print(f"  {sn}!{c.coordinate}: {c.value} → {new_v}")
                    c.value = new_v
                    replacements += 1

wb.save(path)
print(f"\n[fix] 公式范围修复 {replacements} 处, 保存: {path}")
