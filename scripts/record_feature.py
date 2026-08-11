"""
功能开发计划 Excel 记录工具 —— 以后新增功能时同步记录到 docs/reports/2026-08-06_功能开发计划_v3.xlsx。

用法示例：

    # 基本：必填主功能 + 子项目
    python scripts/record_feature.py --group gh-ost --name "v0.4.0 字节压缩"

    # 完整：含状态、风险、人天、说明
    python scripts/record_feature.py \
        --group gh-ost \
        --name "v0.4.0 列压缩支持" \
        --status "待办" \
        --risk "中" \
        --owner "mavis" \
        --days 2 \
        --scenario "大表加列 / 改列类型不再重建" \
        --howto "admin 勾选\"启用列压缩\" → 自动检测 + 调 gh-ost"

    # 从 git 自动拿 commit + 时间
    python scripts/record_feature.py --group gh-ost --name "v0.4.0" --auto-git
"""

import argparse
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path

# Windows PowerShell GBK 输出兼容
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

# Excel 文件路径（相对项目根）
REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = REPO_ROOT / "docs" / "reports" / "2026-08-06_功能开发计划_v3.xlsx"

# 主功能分组（与表格里的分组对应）
# 2026-08-11 新增"提测阶段"分组 —— 收纳 DBA 浏览器验证触发的 bug 修复
VALID_GROUPS = ("平台基础", "钉钉 OA 集成", "gh-ost 无锁 DDL", "收尾", "提测阶段")

# alias 别名（让 --group 输短名也行）
GROUP_ALIASES = {
    "platform": "平台基础", "base": "平台基础", "基建": "平台基础", "基础": "平台基础",
    "oa": "钉钉 OA 集成", "dingtalk": "钉钉 OA 集成", "钉钉": "钉钉 OA 集成", "ding": "钉钉 OA 集成",
    "gh-ost": "gh-ost 无锁 DDL", "ghost": "gh-ost 无锁 DDL", "无锁": "gh-ost 无锁 DDL", "ddl": "gh-ost 无锁 DDL",
    "wrap": "收尾", "cleanup": "收尾", "收尾": "收尾", "尾": "收尾",
    "qa": "提测阶段", "test": "提测阶段", "bug": "提测阶段", "提测": "提测阶段", "测试": "提测阶段",
}

VALID_STATUS = ("已发布", "待办", "TBD", "可选")
VALID_RISK = ("高", "中", "低", "无")
VALID_OWNER = ("mavis", "阿达叔叔", "DBA", "TBD（独立项目）")


def get_git_info():
    """从 git 自动拿 commit hash + 日期。"""
    try:
        sha = subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=REPO_ROOT, stderr=subprocess.DEVNULL, text=True,
        ).strip()
    except Exception:
        sha = "—"
    try:
        # 用 --format=%aI (ISO 8601) 拿完整时间，再 Python 截 YYYY-MM-DD
        # 避免 % 在 Windows shell 被吞
        iso = subprocess.check_output(
            ["git", "log", "-1", "--format=%aI"],
            cwd=REPO_ROOT, stderr=subprocess.DEVNULL, text=True,
        ).strip()
        date = iso[:10]  # 2026-08-06T10:30:00+08:00 → 2026-08-06
        if not date or len(date) < 10:
            date = datetime.now().strftime("%Y-%m-%d")
    except Exception:
        date = datetime.now().strftime("%Y-%m-%d")
    return sha, date


def append_row_to_xlsx(
    xlsx_path: Path,
    group: str,
    name: str,
    when: str,
    pct: float,
    status: str,
    commit: str,
    note: str,
    risk: str,
    owner: str,
    days,
    desc: str,
):
    """追加一行到 v3.xlsx（保留所有样式 + 公式）。"""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} 不存在。先跑一次 build_feature_plan_v3_xlsx.py 建初始表。", file=sys.stderr)
        sys.exit(1)

    try:
        wb = load_workbook(xlsx_path)
    except PermissionError:
        # Excel 进程占用 — 写到一个带时间戳的副本
        from datetime import datetime as _dt
        fallback = xlsx_path.parent / f"{xlsx_path.stem}_{_dt.now().strftime('%H%M%S')}{xlsx_path.suffix}"
        print(f"⚠️  {xlsx_path.name} 被 Excel 进程占用（Permission denied）", file=sys.stderr)
        print(f"   写到副本: {fallback.name}（用户关掉 Excel 后用这个覆盖 v3.xlsx）", file=sys.stderr)
        wb = load_workbook(fallback)
        xlsx_path = fallback
    ws = wb["功能开发计划"]
    new_row = ws.max_row + 1

    # === 样式（与 build_feature_plan_v3_xlsx.py 保持一致）===
    HEADER_FILL = PatternFill("solid", fgColor="2F6F5E")
    GROUP_FILLS = {
        "平台基础":       PatternFill("solid", fgColor="E8F0ED"),
        "钉钉 OA 集成":  PatternFill("solid", fgColor="EAF1F8"),
        "gh-ost 无锁 DDL": PatternFill("solid", fgColor="FBF1E4"),
        "收尾":          PatternFill("solid", fgColor="F2F0EA"),
        "提测阶段":      PatternFill("solid", fgColor="FDEEEE"),  # 浅粉 — bug 修复区
    }
    STATUS_FILLS = {"已发布": PatternFill("solid", fgColor="C6EFCE"), "待办": PatternFill("solid", fgColor="FFEB9C"),
                    "TBD": PatternFill("solid", fgColor="D9D9D9"), "可选": PatternFill("solid", fgColor="D9D9D9")}
    STATUS_FONTS = {"已发布": Font(name="Calibri", size=10, color="006100", bold=True),
                    "待办": Font(name="Calibri", size=10, color="9C5700", bold=True),
                    "TBD": Font(name="Calibri", size=10, color="595959", bold=True),
                    "可选": Font(name="Calibri", size=10, color="595959", bold=True)}
    RISK_FILLS = {"高": PatternFill("solid", fgColor="FFC7CE"), "中": PatternFill("solid", fgColor="FFEB9C"),
                  "低": PatternFill("solid", fgColor="C6EFCE"), "无": PatternFill("solid", fgColor="F2F0EA")}
    RISK_FONTS = {"高": Font(name="Calibri", size=10, color="9C0006", bold=True),
                  "中": Font(name="Calibri", size=10, color="9C5700", bold=True),
                  "低": Font(name="Calibri", size=10, color="006100", bold=True),
                  "无": Font(name="Calibri", size=10, color="595959")}
    OWNER_FILLS = {"mavis": PatternFill("solid", fgColor="E8F0ED"), "阿达叔叔": PatternFill("solid", fgColor="EAF1F8"),
                   "DBA": PatternFill("solid", fgColor="F4E6F0"), "TBD（独立项目）": PatternFill("solid", fgColor="F2F0EA"),
                   "—": PatternFill("solid", fgColor="FFFFFF")}
    OWNER_FONTS = {"mavis": Font(name="Calibri", size=10, color="2F6F5E", bold=True),
                   "阿达叔叔": Font(name="Calibri", size=10, color="1F5B96", bold=True),
                   "DBA": Font(name="Calibri", size=10, color="6B3A86", bold=True),
                   "TBD（独立项目）": Font(name="Calibri", size=10, color="9C5700", bold=True, italic=True)}
    PROGRESS_FILL_100 = PatternFill("solid", fgColor="C6EFCE")
    PROGRESS_FILL_50  = PatternFill("solid", fgColor="FFEB9C")
    PROGRESS_FILL_0   = PatternFill("solid", fgColor="FFC7CE")
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BODY_FONT = Font(name="Calibri", size=10)
    BODY_ALIGN = Alignment(vertical="center", wrap_text=True)
    COMMIT_FONT = Font(name="Consolas", size=9, color="3A4256")
    NOTE_FONT = Font(name="Calibri", size=9, color="595959")
    NOTE_ALIGN = Alignment(vertical="center", wrap_text=True)
    DESC_FONT = Font(name="Calibri", size=9.5, color="14171E")
    DESC_ALIGN = Alignment(vertical="center", wrap_text=True)

    # === 写行 ===
    pct_text = f"{int(pct*100)}%"
    if isinstance(days, (int, float)):
        days_text = f"{days:g}"
    else:
        days_text = str(days)

    values = [group, name, when, pct_text, status, commit, risk, owner, days_text, note, desc]
    for col, val in enumerate(values, 1):
        c = ws.cell(row=new_row, column=col, value=val)
        c.border = BORDER
        c.alignment = BODY_ALIGN
        c.font = BODY_FONT

        if col == 1:
            c.fill = GROUP_FILLS.get(val, PatternFill())
            c.font = Font(name="Calibri", size=10, bold=True, color="14171E")
        elif col == 4:
            if pct >= 0.999: c.fill = PROGRESS_FILL_100
            elif pct > 0:   c.fill = PROGRESS_FILL_50
            else:            c.fill = PROGRESS_FILL_0
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(name="Calibri", size=10, bold=True, color="14171E")
        elif col == 5:
            c.fill = STATUS_FILLS.get(val, PatternFill())
            c.font = STATUS_FONTS.get(val, BODY_FONT)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 6:
            c.font = COMMIT_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 7:
            c.fill = RISK_FILLS.get(val, PatternFill())
            c.font = RISK_FONTS.get(val, BODY_FONT)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 8:
            c.fill = OWNER_FILLS.get(val, PatternFill())
            c.font = OWNER_FONTS.get(val, BODY_FONT)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 9:
            c.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(days, (int, float)):
                c.fill = PatternFill("solid", fgColor="EAF1F8")
                c.font = Font(name="Calibri", size=10, color="1F5B96", bold=True)
            else:
                c.fill = PatternFill("solid", fgColor="F2F0EA")
                c.font = Font(name="Calibri", size=10, color="9C5700", italic=True)
        elif col == 10:
            c.font = NOTE_FONT
            c.alignment = NOTE_ALIGN
        elif col == 11:
            c.fill = PatternFill("solid", fgColor="FAF8F0")
            c.font = DESC_FONT
            c.alignment = DESC_ALIGN

    # 行高
    if desc and len(desc) > 30:
        ws.row_dimensions[new_row].height = 48
    else:
        ws.row_dimensions[new_row].height = 24

    wb.save(xlsx_path)
    return new_row, xlsx_path


def delete_row_by_name(xlsx_path: Path, name: str):
    """按子项目名精确删除一行。"""
    from openpyxl import load_workbook
    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} 不存在", file=sys.stderr)
        sys.exit(1)
    try:
        wb = load_workbook(xlsx_path)
    except PermissionError:
        print(f"ERROR: {xlsx_path.name} 被占用，先关掉 Excel 再删", file=sys.stderr)
        sys.exit(1)
    ws = wb["功能开发计划"]
    deleted = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == name:
            ws.delete_rows(row, 1)
            deleted += 1
            print(f"  删除第 {row} 行: {name}")
    if deleted == 0:
        print(f"  没找到匹配 '{name}' 的行")
    else:
        wb.save(xlsx_path)
        print(f"\n✓ 共删除 {deleted} 行（{xlsx_path}）")


def main():
    parser = argparse.ArgumentParser(
        description="记录新功能到 docs/reports/2026-08-06_功能开发计划_v3.xlsx",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例：
  %(prog)s --group gh-ost --name "v0.4.0 列压缩" --status 待办 --risk 中
  %(prog)s --group gh-ost --name "v0.4.0" --auto-git --scenario "大表加列不再重建" --howto "admin 勾选"
  %(prog)s --list-groups        # 列出 4 个主功能
        """,
    )
    parser.add_argument("--group", help="主功能（平台基础 / 钉钉 OA 集成 / gh-ost 无锁 DDL / 收尾）")
    parser.add_argument("--name", "--sub", dest="name", help="子项目名称")
    parser.add_argument("--status", default="待办", choices=VALID_STATUS, help="状态")
    parser.add_argument("--risk", default="低", choices=VALID_RISK, help="风险等级")
    parser.add_argument("--owner", default="mavis", choices=VALID_OWNER, help="负责人")
    parser.add_argument("--days", help="投入人天（数字或 '待估'）")
    parser.add_argument("--commit", help="关键 commit hash（默认用 git HEAD）")
    parser.add_argument("--when", help="完成时间（YYYY-MM-DD，默认今天）")
    parser.add_argument("--pct", type=float, default=0.0, help="完成进度 0~1（默认 0）")
    parser.add_argument("--note", default="", help="备注（技术细节）")
    parser.add_argument("--scenario", help="场景说明（自动拼到功能说明列）")
    parser.add_argument("--howto", help="使用方法（自动拼到功能说明列）")
    parser.add_argument("--desc", help="完整功能说明（覆盖 --scenario/--howto 拼接）")
    parser.add_argument("--auto-git", action="store_true",
                        help="自动从 git HEAD 拿 commit + 日期")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="xlsx 路径")
    parser.add_argument("--list-groups", action="store_true", help="只列主功能清单")
    parser.add_argument("--dry-run", action="store_true", help="只打印不写")
    parser.add_argument("--delete", metavar="NAME",
                        help="按子项目名删除一行（精确匹配，用于撤销误录）")
    args = parser.parse_args()

    if args.list_groups:
        print("主功能分组（4 个）：")
        for g in VALID_GROUPS:
            print(f"  - {g}")
        sys.exit(0)

    if args.delete:
        delete_row_by_name(Path(args.xlsx), args.delete)
        sys.exit(0)

    if not args.group or not args.name:
        parser.error("--group 和 --name 是必填")

    if args.group not in VALID_GROUPS:
        # 尝试别名匹配
        aliased = GROUP_ALIASES.get(args.group.lower())
        if aliased:
            print(f"  [alias] '{args.group}' → '{aliased}'")
            args.group = aliased
        else:
            print(f"ERROR: --group 必须是 {VALID_GROUPS} 之一（或用 alias：{list(GROUP_ALIASES.keys())[:8]} ...）",
                  file=sys.stderr)
            sys.exit(1)

    # git auto
    commit, when = (args.commit or "—"), (args.when or datetime.now().strftime("%Y-%m-%d"))
    if args.auto_git or (not args.commit and not args.when):
        sha, date = get_git_info()
        if not args.commit:
            commit = sha
        if not args.when:
            when = date

    # days 解析
    days = args.days
    if days is not None:
        try:
            days = float(days)
            if days.is_integer():
                days = int(days)
        except ValueError:
            pass  # 保留字符串（"待估"等）

    print(f"[record_feature] → {args.xlsx}")
    print(f"  主功能: {args.group}")
    print(f"  子项目: {args.name}")
    print(f"  状态: {args.status} ({int(args.pct*100)}%) | 风险: {args.risk} | 负责人: {args.owner} | 人天: {days}")
    print(f"  commit: {commit} | 完成时间: {when}")
    if args.desc:
        print(f"  功能说明: {args.desc[:80]}{'...' if len(args.desc) > 80 else ''}")
    elif args.scenario or args.howto:
        print(f"  功能说明: 【场景】{args.scenario or ''}【使用】{args.howto or ''}")

    if args.dry_run:
        print("\n[DRY-RUN] 不写文件")
        return

    row, final_path = append_row_to_xlsx(
        Path(args.xlsx), args.group, args.name, when, args.pct, args.status,
        commit, args.note, args.risk, args.owner, days, args.desc or "",
    )
    print(f"\n✓ 已写入第 {row} 行")
    print(f"  {final_path}")
    if str(final_path) != args.xlsx:
        print(f"\n⚠️  注意：因 v3.xlsx 被占用，写到了副本 {final_path.name}")
        print(f"   流程：关掉 Excel → 把副本覆盖 v3.xlsx → git commit 副本（可选）")
    print(f"\n提示：打开 Excel 后在【汇总统计】sheet 看公式自动重算；如未生效按 F9 手动刷新。")


if __name__ == "__main__":
    main()
