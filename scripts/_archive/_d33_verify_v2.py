# -*- coding: utf-8 -*-
"""D33 verify v2: 用更准确的检查."""
import paramiko
import base64

DEV = "172.20.2.134"
PWD = "lAqfb8uEmQYsnGNQwIHtGPwukjCz6J"
DEV_BASE = "/opt/archery/prod"

ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
ssh.connect(hostname=DEV, port=22, username="root", password=PWD, timeout=15)

def run(cmd, timeout=30):
    try:
        stdin, stdout, stderr = ssh.exec_command(cmd, timeout=timeout)
        out = stdout.read().decode("utf-8", errors="replace")
        return out
    except Exception as e:
        return f"ERR: {e}"

py_lines = [
    "import os",
    "os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'archery.settings')",
    "import django; django.setup()",
    "from django.test import Client",
    "from sql.models import Users",
    "from django.urls import reverse",
    "",
    "admin = Users.objects.get(username='archery')",
    "c = Client(); c.force_login(admin)",
    "",
    "# 用 reverse() 拿 url",
    "export_url = reverse('ddl_sync:pair_history_export', args=[1])",
    "print('export_url:', export_url)",
    "",
    "# 渲染 page 1",
    "r = c.get('/ddl_sync/pair/1/', HTTP_HOST='172.20.2.134')",
    "print('page 1 status:', r.status_code, 'len:', len(r.content))",
    "html1 = r.content.decode('utf-8', errors='replace')",
    "print('  has export url:', export_url in html1)",
    "print('  has export button class:', 'ddlsync-btn-export' in html1)",
    "print('  has pagination link:', 'history_page=' in html1)",
    "print('  has page link current:', 'ddlsync-page-current' in html1)",
    "",
    "# 渲染 page 2",
    "r2 = c.get('/ddl_sync/pair/1/?history_page=2', HTTP_HOST='172.20.2.134')",
    "print('page 2 status:', r2.status_code, 'len:', len(r2.content))",
    "html2 = r2.content.decode('utf-8', errors='replace')",
    "print('  has export url:', export_url in html2)",
    "",
    "# 测导出 view 直接访问 (用 force_login client)",
    "r3 = c.get(export_url, HTTP_HOST='172.20.2.134')",
    "print('export status:', r3.status_code, 'content-type:', r3.get('Content-Type'))",
    "print('content-disposition:', r3.get('Content-Disposition'))",
    "print('content length:', len(r3.content))",
    "# 写 .xlsx 到 /tmp 验证",
    "with open('/tmp/_d33_test.xlsx', 'wb') as f:",
    "    f.write(r3.content)",
    "print('saved to /tmp/_d33_test.xlsx')",
    "",
    "# 用 openpyxl 验证",
    "try:",
    "    from openpyxl import load_workbook",
    "    wb = load_workbook('/tmp/_d33_test.xlsx')",
    "    ws = wb.active",
    "    print('sheet title:', ws.title)",
    "    print('rows count:', ws.max_row)",
    "    print('header:', [c.value for c in ws[1]])",
    "    if ws.max_row > 1:",
    "        print('row 2:', [c.value for c in ws[2]])",
    "except Exception as e:",
    "    print('openpyxl ERR:', e)",
]
py = '\n'.join(py_lines)
py_b64 = base64.b64encode(py.encode('utf-8')).decode('ascii')
run('echo ' + py_b64 + ' | base64 -d > /tmp/_d33_v3.py')
out = run('cd ' + DEV_BASE + ' && sudo -u archery venv/bin/python manage.py shell < /tmp/_d33_v3.py 2>&1 | head -30 | iconv -f utf-8 -t ascii//IGNORE')
print(out)

ssh.close()
