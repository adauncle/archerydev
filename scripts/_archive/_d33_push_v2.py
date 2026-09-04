# -*- coding: utf-8 -*-
"""D33 push v2: 分页 + Excel 导出."""
import paramiko
import base64

DEV = "172.20.2.134"
PWD = "lAqfb8uEmQYsnGNQwIHtGPwukjCz6J"
DEV_BASE = "/opt/archery/prod"

ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
ssh.connect(hostname=DEV, port=22, username="root", password=PWD, timeout=15)

def run(cmd, timeout=30):
    try:
        stdin, stdout, stderr = ssh.exec_command(cmd, timeout=timeout)
        out = stdout.read().decode("utf-8", errors="replace")
        return out
    except Exception as e:
        return f"ERR: {e}"

print("=" * 60)
print("D33 push v2: 分页 + Excel 导出")
print("=" * 60)

# 1. scp 3 文件
print("\n--- Step 1: scp 3 文件 ---")
files = [
    (r"G:\MiniMax工作空间\archery_dev\sql\extensions\ddl_sync\views\__init__.py", "/tmp/_d33_views.py"),
    (r"G:\MiniMax工作空间\archery_dev\sql\extensions\ddl_sync\urls.py", "/tmp/_d33_urls.py"),
    (r"G:\MiniMax工作空间\archery_dev\sql\extensions\ddl_sync\templates\ddl_sync\pair_detail.html", "/tmp/_d33_pair_detail.html"),
]
sftp = ssh.open_sftp()
for local, remote in files:
    sftp.put(local, remote)
    out = run('ls -la ' + remote + ' 2>&1 | head -1')
    print(f"  {remote}: {out.strip()}")
sftp.close()

# 2. 备份
print("\n--- Step 2: 备份 ---")
run('cp -v ' + DEV_BASE + '/sql/extensions/ddl_sync/views/__init__.py ' + DEV_BASE + '/sql/extensions/ddl_sync/views/__init__.py.bak_d33')
run('cp -v ' + DEV_BASE + '/sql/extensions/ddl_sync/urls.py ' + DEV_BASE + '/sql/extensions/ddl_sync/urls.py.bak_d33')
run('cp -v ' + DEV_BASE + '/sql/extensions/ddl_sync/templates/ddl_sync/pair_detail.html ' + DEV_BASE + '/sql/extensions/ddl_sync/templates/ddl_sync/pair_detail.html.bak_d33')

# 3. 覆盖
print("\n--- Step 3: 覆盖 ---")
run('cp -v /tmp/_d33_views.py ' + DEV_BASE + '/sql/extensions/ddl_sync/views/__init__.py')
run('cp -v /tmp/_d33_urls.py ' + DEV_BASE + '/sql/extensions/ddl_sync/urls.py')
run('cp -v /tmp/_d33_pair_detail.html ' + DEV_BASE + '/sql/extensions/ddl_sync/templates/ddl_sync/pair_detail.html')

# 4. 清 pycache
print("\n--- Step 4: 清 pycache ---")
run('find ' + DEV_BASE + ' -name __pycache__ -type d -exec rm -rf {} + 2>/dev/null')
run('find ' + DEV_BASE + ' -name "*.pyc" -delete 2>/dev/null')
print("pycache cleared")

# 5. kill + 拉新 gunicorn
print("\n--- Step 5: kill + 拉新 gunicorn ---")
run("pkill -9 -f 'gunicorn.*archery.*9003' 2>&1; sleep 2")
import time
out = run('cd ' + DEV_BASE + ' && setsid nohup sudo -u archery venv/bin/gunicorn archery.wsgi:application -w 4 -b 0.0.0.0:9003 --access-logfile - --error-logfile - --timeout 120 </dev/null >/var/log/archery/gunicorn_d33.log 2>&1 & disown')
print(f"gunicorn 拉新: {out.strip()}")
time.sleep(5)
out = run("ps -ef | grep -E 'gunicorn.*9003' | grep -v grep | wc -l")
print(f"gunicorn 进程数 (期望 5): {out.strip()}")

# 6. Django check
print("\n--- Step 6: Django check ---")
out = run('cd ' + DEV_BASE + ' && sudo -u archery venv/bin/python manage.py check 2>&1 | head -10 | iconv -f utf-8 -t ascii//IGNORE')
print(out)

# 7. 验证渲染 + 分页
print("\n--- Step 7: 验证分页 ---")
py_lines = [
    "import os",
    "os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'archery.settings')",
    "import django; django.setup()",
    "from django.test import Client",
    "from sql.models import Users",
    "",
    "admin = Users.objects.get(username='archery')",
    "c = Client(); c.force_login(admin)",
    "",
    "# 第 1 页",
    "r = c.get('/ddl_sync/pair/1/', HTTP_HOST='172.20.2.134')",
    "print('page 1 status:', r.status_code, 'len:', len(r.content))",
    "html1 = r.content.decode('utf-8', errors='replace')",
    "print('  history count mention:', 'xlsx' in html1.lower() or 'history_count' in html1)",
    "print('  has export button:', 'pair_history_export' in html1)",
    "print('  has pagination:', 'history_page=' in html1)",
    "",
    "# 第 2 页",
    "r2 = c.get('/ddl_sync/pair/1/?history_page=2', HTTP_HOST='172.20.2.134')",
    "print('page 2 status:', r2.status_code, 'len:', len(r2.content))",
    "html2 = r2.content.decode('utf-8', errors='replace')",
    "",
    "# 看 history_count",
    "import re",
    "m = re.search(r'\\u540c\\u6b65\\u5386\\u53f2 \\((\\S+)\\s*\\u6761', html1)",
    "if m: print('history count text:', m.group(1))",
]
py = '\n'.join(py_lines)
py_b64 = base64.b64encode(py.encode('utf-8')).decode('ascii')
run('echo ' + py_b64 + ' | base64 -d > /tmp/_d33_v2.py')
out = run('cd ' + DEV_BASE + ' && sudo -u archery venv/bin/python manage.py shell < /tmp/_d33_v2.py 2>&1 | head -30 | iconv -f utf-8 -t ascii//IGNORE')
print(out)

# 8. 测试 Excel 导出
print("\n--- Step 8: 测试 Excel 导出 ---")
out = run("curl -s -I -L 'http://127.0.0.1:9003/ddl_sync/pair/1/history_export/' --max-time 15 2>&1 | head -20")
print(out)
# 实际下载
out = run("curl -s -L -o /tmp/_d33_test.xlsx 'http://127.0.0.1:9003/ddl_sync/pair/1/history_export/' --max-time 15 2>&1")
print(f"download: {out.strip()}")
out = run('ls -la /tmp/_d33_test.xlsx 2>&1')
print(out)
# 验证 .xlsx 格式 (zip 格式)
out = run('file /tmp/_d33_test.xlsx 2>&1')
print(f"file type: {out.strip()}")
# 看 .xlsx 头部 (PK = zip)
out = run('head -c 4 /tmp/_d33_test.xlsx | xxd 2>&1')
print(f"magic bytes: {out.strip()}")
# 看 python 解析
py_test = '''
from openpyxl import load_workbook
wb = load_workbook("/tmp/_d33_test.xlsx")
ws = wb.active
print("sheet title:", ws.title)
print("rows count:", ws.max_row)
print("header:", [c.value for c in ws[1]])
if ws.max_row > 1:
    print("row 2:", [c.value for c in ws[2]])
'''
py_b64_2 = base64.b64encode(py_test.encode('utf-8')).decode('ascii')
run('echo ' + py_b64_2 + ' | base64 -d > /tmp/_d33_xlsx_test.py')
out = run('python3 /tmp/_d33_xlsx_test.py 2>&1')
print(out)

ssh.close()
