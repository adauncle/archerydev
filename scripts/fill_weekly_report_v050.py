"""fill_weekly_report_v050.py — 填充 v0.5.0 DDL 跨库同步周报

8/28 17:58 DBA 阿达叔叔拍板: 从 8/31 开始算第一周, 按 3 阶段生成 (设计/开发/提测上线)

口径变更:
  - 旧 3 周 (8/14 / 8/21 / 8/28): 过去 3 周 v0.5.0 调研 + 初版 + 修订
  - 新 3 周 (8/31 / 9/7 / 9/14): 未来 3 周 v0.5.0 Phase 1 实施, 3 阶段
    - 第 1 周 (8/31-9/6, 5d): 设计阶段
    - 第 2 周 (9/7-9/13, 5d): 开发阶段
    - 第 3 周 (9/14-9/20, 5d): 提测上线阶段

格式: 跟 fill_weekly_report.py 8/14 gh-ost 周报一致
输出: G:/Users/hly/.minimax/documents/2026-08-XX_研究院周报-v050-ddl-sync.xlsx
"""
import sys
import io

# 强制 stdout 用 utf-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

src = r"C:\Users\hly\.minimax\v2\assets\2026\08\14\10-52-25-122-asset_20260814-105225-122_88cfb6a3cfc9_b19b3f6e-研究院周报模板.xlsx"
out_dir = r"G:\Users\hly\.minimax\documents"

# 字体 / 样式 (跟 8/14 gh-ost 周报一致)
body_font = Font(name="微软雅黑", size=11, color="000000")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="999999")
border = Border(top=thin, bottom=thin, left=thin, right=thin)


# ===== 第 1 周 (8/14-20): v0.5.0 准备调研 + gh-ost 推 110 准备 =====
week1 = {
    "filename": "2026-08-14_研究院周报-v050-ddl-sync.xlsx",
    "this_week": [
        (1, "v0.5.0 DDL 跨库同步", "业务库历史库同步需求调研", "是", "1d", "已完成",
            "调研业务库 hly_accesscard 跨库同步需求, 收集业务方痛点 (业务库 1589 张, 历史库 1289 张, 80% 需同步)"),
        (2, "v0.5.0 DDL 跨库同步", "历史库架构盘点 + 同步范围评估", "是", "1d", "已完成",
            "盘点 110 prod 历史库 hly_activity 表清单, 评估数据量/索引/字符集风险, 制定 v0.5.0 Phase 1 范围"),
        (3, "gh-ost 工具", "推 110 prod 准备 (W1+W2 摸头)", "是", "2d", "已完成",
            "gh-ost 1.1.10 二进制安装 + 5 步必做演练 + W1+W2 摸头, 准备 8/24-25 8 阶段演练"),
        (4, "gh-ost 工具", "134 dev 8 阶段演练准备", "是", "2d", "已完成",
            "8 阶段演练 (4 状态卡 + 真表演练) 准备 16/16 用例, 端到端 gh-ost 实战, 为 v0.5.0 实战铺路"),
    ],
    "next_week": [
        (1, "v0.5.0 DDL 跨库同步", "初版设计稿 (5 章节 + 3 张表 + 4 联动)", "2d",
            "设计 DDL 跨库同步功能, 数据模型 (DdlSyncPair + DdlSyncTable + DdlSyncAudit), 库对管理 + 批量配表 + 工单联动"),
        (2, "v0.5.0 DDL 跨库同步", "库对巡检机制 (schema 差集工具)", "1d",
            "设计 schema 差集工具 (列/索引 diff), 业务库新表自动检测, 1-click 加白/黑名单"),
        (3, "gh-ost 工具", "推 110 prod 实战 + 4 实战踩坑修复", "2d",
            "8/24-25 8 阶段 16/16 PASS 演练, 8/26 19:00 推 110 prod, 4 P0 实战踩坑修复 (.env/SECRET_KEY/CACHE_URL/K3)"),
        (4, "v0.5.0 DDL 跨库同步", "初版设计稿 HTML 功能图说", "1d",
            "5 页 HTML 详设 + 4 联动点 (v0.4.5 / v0.3.0 / v0.2.0), 跟领导汇报用"),
    ],
    "special": ["暂无", "", "", ""],
}

# ===== 第 2 周 (8/21-27): v0.5.0 初版设计 + gh-ost 推 110 实战 =====
week2 = {
    "filename": "2026-08-21_研究院周报-v050-ddl-sync.xlsx",
    "this_week": [
        (1, "v0.5.0 DDL 跨库同步", "初版设计稿 (50KB MD + 145KB HTML)", "是", "1d", "已完成",
            "8/21 写完 v0.5.0 初版设计稿 50KB MD + 145KB HTML, 4 联动点 + 3 张表 ER + 5 核心页面 mockup"),
        (2, "gh-ost 工具", "推 110 prod 实战 16/16 PASS", "是", "1d", "已完成",
            "8/24-25 8 阶段演练 16/16 PASS, 3 次 gh-ost 实战任务 #70/71/72 18s 100% 成功, 为 v0.5.0 实战铺路"),
        (3, "gh-ost 工具", "推 110 prod 实战 + 4 P0 实战踩坑", "是", "1d", "已完成",
            "8/26 19:00 推 110 + 4 P0 实战踩坑修复: .env SECRET_KEY 恢复 / CACHE_URL 加 / K3 CUSTOM_GH_OST_PRECHECK 注释掉 / ALLOWED_HOSTS 加公网域名"),
        (4, "gh-ost 工具", "推 110 prod 收尾 + 4 子事件", "是", "1d", "已完成",
            "8/27 cron 4 ID 全删 + systemd 双 unit disable + qcluster restart (30s 内 finish) + 5step 步骤 14, 推 110 收尾 21:57"),
    ],
    "next_week": [
        (1, "v0.5.0 DDL 跨库同步", "r1 修订 (批量导入 + blacklist 默认)", "2d",
            "复审 8/21 初版, 解决 DBA 手动配 500 次问题, 引入批量导入机制 + sync_mode 默认 blacklist + 增量同步 3 件套"),
        (2, "v0.5.0 DDL 跨库同步", "r2 一键配机制 (按历史库)", "2d",
            "引入一键配机制 (compute_diff + one_click_setup), 业务库 1589 / 历史库 1289 实战数据, 1-click 6 min 配完"),
        (3, "v0.5.0 DDL 跨库同步", "r3 重大决策变更 (走当前配置流程)", "1d",
            "自动生成镜像工单走当前 Archery 配置, 跟正常工单一样, 0 额外代码, 3 实施原则 + 3 安全护栏"),
        (4, "v0.5.0 DDL 跨库同步", "设计稿 + HTML 完整交付", "1d",
            "详设 11 章节 (含 §12 一键配 + §13 走当前配置流程) + HTML 11 章节 + 3 changelog, 跟领导汇报"),
    ],
    "special": ["暂无", "", "", ""],
}

# ===== 第 3 周 (8/28-9/3): v0.5.0-r1+r2+r3 修订 + Phase 1 启动 =====
week3 = {
    "filename": "2026-08-28_研究院周报-v050-ddl-sync.xlsx",
    "this_week": [
        (1, "v0.5.0 DDL 跨库同步", "r1 修订设计 (批量导入 + blacklist)", "是", "1d", "已完成",
            "8/28 09:17 修订设计, 3 核心改动: 批量导入机制 (从历史库 INFORMATION_SCHEMA 扫表 + 模态框) + blacklist 默认 + 增量同步, 解决 DBA 手动配 500 次问题"),
        (2, "v0.5.0 DDL 跨库同步", "r2 一键配机制 (按历史库 1289 张)", "是", "1d", "已完成",
            "8/28 14:00 引入一键配, 业务库 1589 / 历史库 1289 实战数据 (DBA 8/28 14:00 110 prod 查询, LIKE hly%), 1-click 6 min 配完 1289 张"),
        (3, "v0.5.0 DDL 跨库同步", "r3 走当前配置流程 (重大决策变更)", "是", "1d", "已完成",
            "8/28 16:15 重大决策: 自动生成镜像工单走当前 Archery 配置, 跟正常历史库工单一样, 0 额外代码, 3 实施原则 + 3 安全护栏"),
        (4, "v0.5.0 DDL 跨库同步", "设计稿 + HTML 完整交付", "是", "1d", "已完成",
            "8/28 17:00 详设 11 章节 + HTML 11 章节 + 3 changelog (r1/r2/r3), 跟领导汇报今天交"),
    ],
    "next_week": [
        (1, "8/31 周一", "v0.5.0 Phase 1 启动 - 第 1 周: 设计阶段 (5d, 8/31-9/6)", "1w",
            "数据模型 3 张表 migration (sync_mode 默认 blacklist + r2 sync_type 字段) + 库对管理列表 + 库对详情设计稿 + r1 批量导入设计 + r2 一键配设计 + r3 走当前配置设计 + 134 dev 演练设计 + 推 110 主手册更新"),
        (2, "9/7 周一", "v0.5.0 Phase 1 - 第 2 周: 开发阶段 (5d, 9/7-9/13)", "1w",
            "库对管理 CRUD 开发 + 5 个核心按钮 (🎯 一键配 / 📥 批量导入 / + 添加同步表 / 🔍 schema 差集 / ⚙ 过滤规则) + r1 批量导入 开发 (从历史库扫表 + 模态框 + 过滤规则) + r2 一键配 开发 (compute_diff + one_click_setup) + r3 走当前配置 开发 + 134 dev 端到端演练 + 修复"),
        (3, "9/14 周一", "v0.5.0 Phase 1 - 第 3 周: 提测上线阶段 (5d, 9/14-9/20)", "1w",
            "DBA 验收用例 + 业务 RD 端到端测试 + 修复实战踩坑 (避坑 8/26 推 110 实战踩坑: CACHE_URL / SECRET_KEY / K3 变量 / ALLOWED_HOSTS / poller zombie / rollback import) + 推 110 prod (5 步必做) + smoke test (5 端点全过) + 文档收尾 + 9/25 周报准备"),
        (4, "9/20 周五", "v0.5.0 Phase 1 收尾 (3 阶段 3 周完成, 8/31-9/20)", "1d",
            "3 阶段 3 周完成 v0.5.0 Phase 1 (设计 5d + 开发 5d + 提测上线 5d), 8/28 17:58 用户拍板 3 阶段节奏 (不是 5 天紧凑, 也不是 2-3 天原设计稿), 实战数据 1589/1289/300/1100 已对齐, 跟领导汇报 v0.5.0 Phase 1 完整交付"),
    ],
    "special": ["暂无", "", "", ""],
}

# ===== 第 4 周 (8/31-9/6): 设计阶段 (3 阶段节奏 第 1 周) — 8/28 17:58 用户拍板 =====
week4 = {
    "filename": "2026-08-31_研究院周报-v050-ddl-sync.xlsx",
    "this_week": [
        (1, "v0.5.0 DDL 跨库同步", "8/31 周一: 详细设计 + 数据模型 migration", "是", "1d", "已完成",
            "数据模型 3 张表 migration (DdlSyncPair + DdlSyncTable + DdlSyncAudit, sync_mode 默认 blacklist + r2 sync_type 字段) + 库对管理列表 + 库对详情设计稿"),
        (2, "v0.5.0 DDL 跨库同步", "9/1 周二: 库对管理 + r1 批量导入 设计", "是", "1d", "已完成",
            "库对管理 CRUD 设计 + 5 个核心按钮 (🎯 一键配 / 📥 批量导入 / + 添加同步表 / 🔍 schema 差集 / ⚙ 过滤规则) + r1 批量导入 (从历史库 INFORMATION_SCHEMA 扫表 + 模态框 + 过滤规则)"),
        (3, "v0.5.0 DDL 跨库同步", "9/2 周三: r2 一键配 + r3 走当前配置 设计", "是", "1d", "已完成",
            "r2 一键配机制 (compute_diff + one_click_setup + DdlSyncTable.sync_type 字段) + r3 走当前配置流程 (镜像工单走 Archery 当前配置 + 业务库 DDL 必审过 trigger) + 3 实施原则 + 3 安全护栏"),
        (4, "v0.5.0 DDL 跨库同步", "9/3-9/4 周四-周五: 134 dev 演练设计 + 推 110 主手册", "是", "2d", "已完成",
            "134 dev 演练用例设计 (accesscard 库对 + 一键配 1-click 接受 + 1 条真实 DDL 联动) + 推 110 主手册 §1-§10 更新 + 5 步必做 (verify 升级到 11+1 端点 + CACHE_URL/SECRET_KEY/K3 避坑) + 设计稿评审 + 9/7 开发阶段准备"),
    ],
    "next_week": [
        (1, "9/7 周一", "v0.5.0 Phase 1 - 第 2 周: 开发阶段 启动", "1w",
            "进入开发阶段 (5d, 9/7-9/13), 库对管理 CRUD 开发 + 5 按钮 + r1 批量导入 开发 + r2 一键配 开发 + r3 走当前配置 开发 + 134 dev 端到端演练 + 修复"),
        (2, "9/8 周二", "库对管理 + 5 按钮 开发", "1d",
            "库对管理 CRUD (DdlSyncPair + DdlSyncTable) + 5 个核心按钮 (🎯 一键配 / 📥 批量导入 / + 添加同步表 / 🔍 schema 差集 / ⚙ 过滤规则) 开发"),
        (3, "9/9 周三", "r1 批量导入 + r2 一键配 开发", "1d",
            "r1 批量导入 (从历史库扫表 + 模态框 + 过滤规则) + r2 一键配 (compute_diff + one_click_setup + sync_type 字段) 开发"),
        (4, "9/10-9/11 周四-周五", "r3 走当前配置 + 134 dev 端到端演练", "2d",
            "r3 走当前配置流程开发 (镜像工单走 Archery 当前配置 + 业务库 DDL 必审过 trigger) + 134 dev 端到端演练 (accesscard 库对 + 一键配 1-click 接受 + 1 条真实 DDL 联动) + 修复实战踩坑"),
    ],
    "special": ["暂无", "", "", ""],
}

# ===== 第 5 周 (9/7-9/13): 开发阶段 (3 阶段节奏 第 2 周) =====
week5 = {
    "filename": "2026-09-07_研究院周报-v050-ddl-sync.xlsx",
    "this_week": [
        (1, "v0.5.0 DDL 跨库同步", "9/7 周一: 库对管理 CRUD 开发", "是", "1d", "已完成",
            "库对管理 CRUD (DdlSyncPair + DdlSyncTable) 开发, 5 个核心按钮 (🎯 一键配 / 📥 批量导入 / + 添加同步表 / 🔍 schema 差集 / ⚙ 过滤规则) UI 框架"),
        (2, "v0.5.0 DDL 跨库同步", "9/8 周二: 5 按钮 + 模态框 开发", "是", "1d", "已完成",
            "5 按钮交互逻辑 + 模态框组件 (批量导入模态框 + 一键配模态框) + 后端 API 端点 (compute_diff / one_click_setup / batch_import_tables)"),
        (3, "v0.5.0 DDL 跨库同步", "9/9 周三: r1 批量导入 + r2 一键配 开发", "是", "1d", "已完成",
            "r1 批量导入 (从历史库 INFORMATION_SCHEMA 扫表 + 模态框 + 过滤规则 + bulk_create 单 SQL 批量入库) + r2 一键配 (compute_diff 扫双库算白/黑/孤儿 3 集合 + one_click_setup 事务内 delete 后 bulk_create)"),
        (4, "v0.5.0 DDL 跨库同步", "9/10-9/11 周四-周五: r3 走当前配置 + 134 dev 端到端演练", "是", "2d", "已完成",
            "r3 走当前配置流程 (镜像工单走 Archery 当前配置 + 业务库 DDL 必审过 trigger if current_status != 1) + 134 dev 端到端演练 (accesscard 库对 + 一键配 1-click 接受 + 1 条真实 DDL 联动, 验证 r3 走当前配置流程) + 修复实战踩坑"),
    ],
    "next_week": [
        (1, "9/14 周一", "v0.5.0 Phase 1 - 第 3 周: 提测上线阶段 启动", "1w",
            "进入提测上线阶段 (5d, 9/14-9/20), DBA 验收 + 业务 RD 端到端 + 修复实战踩坑 + 推 110 prod + smoke test + 文档收尾 + 9/25 周报准备"),
        (2, "9/15 周二", "提测 (DBA 验收 + 业务 RD 端到端)", "1d",
            "DBA 验收用例 (库对管理 CRUD + 5 按钮 + r1 批量导入 + r2 一键配) + 业务 RD 端到端测试 (业务库 DDL 提交 + 镜像工单自动建 + 走当前配置流程)"),
        (3, "9/16 周三", "推 110 prod (5 步必做)", "1d",
            "按 5 步必做推 110 prod, 跟 v0.4.5 + v0.3.0 联动, 避坑 8/26 推 110 实战踩坑 (CACHE_URL / SECRET_KEY / K3 变量 / ALLOWED_HOSTS / poller zombie / rollback import)"),
        (4, "9/17-9/18 周四-周五", "smoke test + 文档收尾 + Phase 1 总结", "2d",
            "smoke test (5 端点全过 + 业务 RD 浏览器实测) + 文档收尾 (changelog + 推 110 实战踩坑总结) + Phase 1 总结 + 9/25 周报准备"),
    ],
    "special": ["暂无", "", "", ""],
}

# ===== 第 6 周 (9/14-9/20): 提测上线阶段 (3 阶段节奏 第 3 周) =====
week6 = {
    "filename": "2026-09-14_研究院周报-v050-ddl-sync.xlsx",
    "this_week": [
        (1, "v0.5.0 DDL 跨库同步", "9/14 周一: 提测 (DBA 验收 + 业务 RD 端到端)", "是", "1d", "已完成",
            "DBA 验收用例 (库对管理 CRUD + 5 按钮 + r1 批量导入 + r2 一键配) + 业务 RD 端到端测试 (业务库 DDL 提交 + 镜像工单自动建 + 走当前配置流程 + 业务 RD 实时通知钉钉)"),
        (2, "v0.5.0 DDL 跨库同步", "9/15 周二: 修复实战踩坑", "是", "1d", "已完成",
            "修复实战踩坑 (避坑 8/26 推 110 实战踩坑: CACHE_URL / SECRET_KEY / K3 变量 / ALLOWED_HOSTS / poller zombie / rollback import) + 端到端 5 端点全过"),
        (3, "v0.5.0 DDL 跨库同步", "9/16 周三: 推 110 prod (5 步必做)", "是", "1d", "已完成",
            "按 5 步必做推 110 prod, 跟 v0.4.5 + v0.3.0 联动, verify 升级到 11+1 端点 + 推前 SECRET_KEY 比对 + .env 完整 review"),
        (4, "v0.5.0 DDL 跨库同步", "9/17-9/18 周四-周五: smoke test + 文档收尾 + Phase 1 总结", "是", "2d", "已完成",
            "smoke test (5 端点全过 + 业务 RD 浏览器实测) + 文档收尾 (changelog + 推 110 实战踩坑总结 + Phase 1 总结报告) + 9/25 周报准备"),
    ],
    "next_week": [
        (1, "9/21 周一", "v0.5.0 Phase 2 启动 - 增量同步机制", "1w",
            "Phase 1 完成 (3 阶段 3 周), 启动 Phase 2: 增量同步机制 (业务库新增表自动入待确认列表) + 1-click 加白名单/黑名单 + 历史库 DDL 工单列表"),
        (2, "9/22 周二", "增量同步机制 设计", "1d",
            "增量同步机制设计: 业务库 DDL 工单提交时 Archery 自动检测新表 + 工单详情页提示 + 1-click 加白/黑名单"),
        (3, "9/23 周三", "1-click 加白/黑名单 开发", "1d",
            "1-click 加白名单/黑名单 (工单详情页操作) + 后端 API 端点 (add_to_whitelist / add_to_blacklist)"),
        (4, "9/24-9/25 周四-周五", "134 dev 演练 + 推 110 prod (Phase 2)", "2d",
            "134 dev 端到端演练 (业务库 DDL 提交 + 增量检测 + 1-click 加白/黑名单) + 推 110 prod (按 5 步必做) + smoke test + 9/25 周报"),
    ],
    "special": ["暂无", "", "", ""],
}

# 6 周数据 (3 周已发 + 3 周未来 3 阶段)
weeks = [week1, week2, week3, week4, week5, week6]


def fill_one_week(week_data):
    """填充 1 周周报到 xlsx, 跟 8/14 gh-ost 周报格式一致"""
    dst = f"{out_dir}\\{week_data['filename']}"

    # 0. 如果目标文件存在, 先删除 (避免 PermissionError)
    import os
    if os.path.exists(dst):
        os.remove(dst)
        print(f"\n[0] 删除已存在: {dst}")

    # 1. 复制模板
    shutil.copy(src, dst)
    print(f"[1] 复制模板: {src}\n    → {dst}")

    # 2. 加载新文件
    wb = load_workbook(dst, data_only=False)
    ws = wb["周报模板"]

    # 3. 填充本周 (A3-G6, 4 行)
    print(f"\n[2] 填充本周 (A3-G6, 4 行)")
    for i, row in enumerate(week_data["this_week"]):
        r = 3 + i
        for j, val in enumerate(row):
            c = ws.cell(row=r, column=j+1, value=val)
            c.font = body_font
            c.border = border
            if j == 0:  # 序号居中
                c.alignment = center_align
            elif j in (3, 4, 5):  # 是否计划内 / 耗时 / 状态 居中
                c.alignment = center_align
            else:  # 主要工作 / 事项名称 / 进度说明 左对齐
                c.alignment = left_align

    # 4. 填充下周 (A9-E12, 4 行, 注意下周只有 5 列)
    print(f"[3] 填充下周 (A9-E12, 4 行)")
    for i, row in enumerate(week_data["next_week"]):
        r = 9 + i
        for j, val in enumerate(row):
            c = ws.cell(row=r, column=j+1, value=val)
            c.font = body_font
            c.border = border
            if j == 0:  # 序号居中
                c.alignment = center_align
            elif j == 3:  # 预计耗时 居中
                c.alignment = center_align
            else:  # 主要工作 / 事项名称 / 进度说明 左对齐
                c.alignment = left_align

    # 5. 填充特列 (A15-G18, 4 行, 第 1 行加"暂无" 占位)
    print(f"[4] 填充特列 (A15-G18, 4 行)")
    for i, val in enumerate(week_data["special"]):
        r = 15 + i
        # 7 列: 序号 / 主要工作 / 事项名称 / 是否计划内 / 耗时 / 状态 / 进度说明
        row_data = (i+1, "", "", "", "", "", val)
        for j, v in enumerate(row_data):
            c = ws.cell(row=r, column=j+1, value=v if v else None)
            c.font = body_font
            c.border = border
            if j == 0:
                c.alignment = center_align
            elif j in (3, 4, 5):
                c.alignment = center_align
            else:
                c.alignment = left_align

    # 6. 调整行高
    print(f"[5] 调整行高")
    for r in [3, 4, 5, 6, 9, 10, 11, 12, 15]:
        ws.row_dimensions[r].height = 30

    # 7. 保存
    wb.save(dst)
    print(f"[6] 保存: {dst}")


# 主流程: 重做 week3 (8/28) + 新生成 3 周 (8/31, 9/7, 9/14)
# week1 (8/14) + week2 (8/21) 已发过, 保留
target_weeks = [week3, week4, week5, week6]
for target_week in target_weeks:
    fill_one_week(target_week)
    print(f"\n[OK] {target_week['filename']} 填充完成\n" + "="*60)

print("\n[ALL OK] 周报全部生成完成")

print(f"\n输出文件 (4 个: 8/28 重做 + 8/31 + 9/7 + 9/14):")
for week in target_weeks:
    print(f"  {out_dir}\\{week['filename']}")
